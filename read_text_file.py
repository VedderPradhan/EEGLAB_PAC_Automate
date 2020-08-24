from pywinauto import application
from pywinauto.keyboard import send_keys
import time
import pyautogui as pg
import openpyxl
from datetime import datetime
import shutil
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

params = [['0.5 1','80 200'], ['0.5 1','200 300'],
          ['3 4','80 200'] , ['3 4','200 300']]
#ampRange = ['80 200', '200 300']
edfFiles = ['No 1 for MI.edf', 'No 2 for MI.edf', 'No 3 for MI.edf',
            'No 4 for MI.edf', 'No 5 for MI.edf', 'No 6 for MI.edf',
            'No 7 for MI.edf', 'No 8 for MI.edf', 'No 9 for MI.edf',
            'No 10 for MI.edf']
workPath = input("Enter path to work on PAC computation: ")

#Create an Excel file to save the results
now = datetime.now()
dt_string = now.strftime("%d-%m-%Y_%H-%M-%S")
fileName = "Results_" + dt_string + ".xlsx"
wb_new = openpyxl.Workbook()
wb_new.save(fileName)

#fileName = 'PutData.xlsx'
lastColumn = 1 #To start with the first column
firstIteration = True #The computed PAC for the first edf File (all parameters)
count = 0 #Guide counter for Row Record indexing 
#RowRec = [0] * len(phaseFreqs) #Tracking the Row Entries
rowIndex = 0 #For storing RowRec[count] in the Iterations excluding the first iteration 
#ampRangeIndex = 0
IsFirstEdfFile = True
rowToEnterData = 1

#Windows and Dialogs definition
app = application.Application().connect(path = r"C:\Program Files\MATLAB\R2019b\bin\win64\MATLAB.exe")
dlg = app.window(title_re = ".*MATLAB*")
dlg2 = app.window(title_re = ".*pac_pop_main.*")
dlg3 = app.window(title_re = ".*pac_pop_statsSetUp.*")

#copy .m file to the target path
shutil.copy("ScriptTest.m", workPath)
with open(workPath + r"\ScriptTest.m" , 'r') as file:
    filedata2 = file.read()

filedata2 = filedata2.replace("path", workPath)

with open(workPath + r"\ScriptTest.m", 'w') as file:
    file.write(filedata2)

#Main Part
for edfFile in edfFiles:
    #Change the edf file name in the .m file
    logging.debug('Entered the edfFile loop' + edfFile)
    with open(workPath + "\ScriptTest.m", 'r') as file:
        filedata = file.read()

    if IsFirstEdfFile == True:
        IsFirstEdfFile = False
    else:
        filedata = filedata.replace(edfFiles[edfFiles.index(edfFile)-1], edfFile)
        print("File changed")

    with open(workPath + "\ScriptTest.m", 'w') as file:
        file.write(filedata)
    
    for param in params:        
        #Activate Matlab window and run ScriptTest.m
        dlg.set_focus()
        time.sleep(0.3)
        pg.keyDown('ctrl')
        pg.press('0')
        pg.keyUp('ctrl')
        pg.write('ScriptTest')
        pg.press('enter')
        

        #Wait for pac_pop_main window to be opened from the ScriptTest.m
        app.wait_cpu_usage_lower(threshold=2, timeout=None, usage_interval=None)
        time.sleep(2.5)
        waitVar = dlg2.wait("exists enabled visible ready active", timeout=30, retry_interval=20)
        print("PAC pop main Window Opened")
        

        #Enter the Parameters in the pac_pop_man window and press OK button
        pg.press('tab')
        pg.write(param[0])
        pg.press('tab')       
        pg.write(param[1])
        pg.press('tab')
        pg.write('5')
        pg.press( 'tab', presses=4, interval=0.3)
        pg.write('0.05')
        pg.press('tab')
        pg.write('500')
        pg.press('tab')
        pg.write('18')
        pg.press( 'tab', presses=2, interval=0.3)
        time.sleep(0.3)
        pg.press('space')
        print('Parameter Entered')            


        #Press OK button when pac_pop_statsSetUp window appears
        app.wait_cpu_usage_lower(threshold=2, timeout=None, usage_interval=None)
        time.sleep(2.5)
        waitVar = dlg3.wait("exists enabled visible ready active", timeout=30, retry_interval=20)
        pg.press( 'tab', presses=7, interval=0.3)
        pg.press('space')
        time.sleep(2)
        
        #open the txt file containing EEG.pac.mi value and read
        textContent = 0
        logging.debug("textContent Value (after setting to 0): " + str(textContent))
        f = open(r"C:\Users\Samantha\Desktop\Sammy - Ope ECoG\Motohashi Yuunosuke 24.8.2020 R temporal ganglioglioma or DNET\Post\saveTest.txt")
        textContent = f.readlines()
        logging.debug("saveTest.txt read")
        print(textContent)
        f.close()
        

        #Write data to Excel
        wb = openpyxl.load_workbook(fileName) #xlsx should exist
        ##wb = openpyxl.Workbook()
        ws1 = wb.active

        #Record the number of channels
        channels = len(textContent)
        rowsReqForEachCompute = channels + 3
        
        ws1.cell(row=rowToEnterData , column=lastColumn, value=edfFile)
        rowToEnterData = rowToEnterData + 1
        
        for text in textContent:
            
            ws1.cell(row=rowToEnterData, column = lastColumn, value = float(text))
            rowToEnterData = rowToEnterData + 1
            print(float(text))
        
        rowToEnterData = rowToEnterData + 3
        wb.save(fileName)
        
    lastColumn = lastColumn + 1
    rowToEnterData = 1
    
    



### Open excel and your workbook
##col = 2 # column B
##excel=win32com.client.Dispatch("Excel.Application")
##excel.Visible=True # Note: set to false when scripting, only True for this example
##wb=excel.Workbooks.Open('PutData.xlsx')
##ws = wb.Worksheets('Sheet1')
##
###Write text contents to column range
##ws.Range(ws.Cells(col ,1),ws.Cells(col,len(text_contents))).Value = text_contents
##
###Save the workbook and quit
##wb.Close(True)
##excel.Application.Quit() 
